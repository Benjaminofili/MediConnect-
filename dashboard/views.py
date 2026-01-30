import os
import boto3
from botocore.config import Config
from django.http import HttpResponseRedirect, HttpResponseForbidden
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.http import JsonResponse
from django.shortcuts import render, redirect,get_object_or_404
from django.contrib.auth import login, logout, authenticate, get_user_model,update_session_auth_hash
from django.contrib import messages
from django.db import models
from django.db.models import Count, Sum, Q, Max
from django.utils import timezone
from datetime import datetime, timedelta,date
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.views.generic import ListView, DetailView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes, force_str
from django.contrib.auth.tokens import default_token_generator
from django.views.decorators.csrf import ensure_csrf_cookie
from django.conf import settings
from django.urls import reverse
from .decorators import doctor_required, patient_required, redirect_authenticated_user
from accounts.models import PatientProfile, DoctorProfile
from django.contrib.auth.decorators import login_required
from accounts.serializers import (
    PatientRegistrationSerializer, 
    DoctorRegistrationSerializer,
    PatientProfileSerializer,
    DoctorProfileSerializer
)
from accounts.models import PatientProfile, DoctorProfile
from appointments.models import Appointment
from doctors.models import *
from consultations.models import Consultation, Prescription
from notifications.services import EmailService
from records.models import HealthProfile, MedicalHistory, MedicalDocument
from records.forms import HealthProfileForm, MedicalHistoryForm, MedicalDocumentForm

User = get_user_model()


# ============================================
# HELPER FUNCTIONS
# ============================================

def get_verified_doctors():
    """Get all verified doctors for appointment forms."""
    from accounts.models import DoctorProfile
    return DoctorProfile.objects.filter(
        verification_status='verified',
        user__is_active=True
    ).select_related('user', 'specialization').order_by('user__first_name')

# ============== AUTH VIEWS ==============

# dashboard/views.py
@redirect_authenticated_user
@ensure_csrf_cookie
def login_page(request):
    """Handle user login with email verification check."""
    # Clear non-auth related messages on login page load
    # This prevents document/profile messages from showing on login
    if request.method == 'GET':
        # Get all messages and filter out non-auth related ones
        storage = messages.get_messages(request)
        auth_keywords = ['logout', 'welcome', 'verify', 'verification', 'password', 'email']
        non_auth_keywords = ['document', 'profile', 'medical history', 'uploaded', 'deleted', 'updated successfully']
        
        messages_to_keep = []
        for message in storage:
            message_str = str(message).lower()
            # Keep error/warning messages, or messages with auth keywords
            is_auth_related = any(keyword in message_str for keyword in auth_keywords)
            is_non_auth = any(keyword in message_str for keyword in non_auth_keywords)
            
            if message.tags in ['error', 'warning'] or (is_auth_related and not is_non_auth):
                messages_to_keep.append((message.tags, str(message)))
        
        # Mark all messages as consumed
        storage.used = True
        
        # Re-add only auth-related messages
        for tag, msg_text in messages_to_keep:
            if tag == 'error':
                messages.error(request, msg_text)
            elif tag == 'warning':
                messages.warning(request, msg_text)
            elif tag == 'success':
                messages.success(request, msg_text)
            else:
                messages.info(request, msg_text)
    
    if request.method == 'POST':
        email = request.POST.get('email', '').strip().lower()
        password = request.POST.get('password')
        remember_me = request.POST.get('remember_me')
        
        print(f"DEBUG: Login attempt for: {email}")
        
        # First check if user exists
        try:
            user_exists = User.objects.get(email=email)
            print(f"DEBUG: User found - email_verified: {user_exists.email_verified}")
        except User.DoesNotExist:
            print(f"DEBUG: User not found: {email}")
            user_exists = None
        
        user = authenticate(request, email=email, password=password)
        print(f"DEBUG: Authentication result: {user}")
        
        if user is not None:
            # Check if email is verified
            if not user.email_verified:
                print(f"DEBUG: Email not verified for {email}")
                messages.warning(
                    request, 
                    f'Please verify your email before logging in. '
                    f'<a href="/resend-verification/" class="alert-link">Resend verification email</a>'
                )
                return render(request, 'dashboard/auth/login.html', {'email': email})
            
            # Check doctor verification status
            if user.user_type == 'doctor':
                try:
                    doctor_profile = user.doctor_profile
                    if doctor_profile.verification_status == 'rejected':
                        messages.error(request, 'Your doctor profile has been rejected.')
                        return render(request, 'dashboard/auth/login.html', {'email': email})
                except DoctorProfile.DoesNotExist:
                    pass
            
            # Login user
            login(request, user)
            print(f"DEBUG: User logged in successfully: {user.email}")
            
            # Handle remember me
            if not remember_me:
                request.session.set_expiry(0)
            
            messages.success(request, f'Welcome back, {user.first_name}!')
            
            # Check for 'next' parameter to redirect back to previous page
            next_url = request.GET.get('next') or request.POST.get('next')
            if next_url:
                return redirect(next_url)
            
            # Redirect based on user type
            if user.user_type == 'doctor':
                return redirect('dashboard:doctor_dashboard')
            elif user.user_type == 'patient':
                return redirect('dashboard:patient_dashboard')
            else:
                return redirect('admin:index')
        else:
            print(f"DEBUG: Authentication failed for {email}")
            messages.error(request, 'Invalid email or password.')
    
    return render(request, 'dashboard/auth/login.html')

@redirect_authenticated_user
def register_page(request):
    """Registration choice page - doctor or patient."""
    return render(request, 'dashboard/auth/register_choice.html')

@redirect_authenticated_user
def verification_sent_page(request):
    """Show verification email sent message."""
    return render(request, 'dashboard/auth/verification_sent.html')


@redirect_authenticated_user
def verify_email(request, uidb64, token):
    """Verify user email from link."""
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = User.objects.get(pk=uid)
        print(f"DEBUG: Found user: {user.email}, email_verified: {user.email_verified}")
    except (TypeError, ValueError, OverflowError, User.DoesNotExist) as e:
        print(f"DEBUG: Error finding user: {e}")
        user = None
    
    if user is None:
        messages.error(request, 'Invalid verification link.')
        return redirect('dashboard:login')
    
    if user.email_verified:
        messages.info(request, 'Your email is already verified. Please log in.')
        return redirect('dashboard:login')
    
    # Check token
    token_valid = default_token_generator.check_token(user, token)
    print(f"DEBUG: Token valid: {token_valid}")
    
    if token_valid:
        user.email_verified = True
        user.save(update_fields=['email_verified'])
        print(f"DEBUG: User {user.email} email_verified updated to: {user.email_verified}")
        
        messages.success(request, 'Email verified successfully! You can now log in.')
        return redirect('dashboard:login')
    else:
        messages.error(request, 'This verification link is invalid or has expired.')
        return redirect('dashboard:resend_verification')

@redirect_authenticated_user
def resend_verification(request):
    """Resend verification email."""
    if request.method == 'POST':
        email = request.POST.get('email', '').strip().lower()
        
        try:
            user = User.objects.get(email=email)
            
            if user.email_verified:
                messages.info(request, 'Your email is already verified. Please log in.')
                return redirect('dashboard:login')
            
            # Generate new token
            token = default_token_generator.make_token(user)
            uid = urlsafe_base64_encode(force_bytes(user.pk))
            
            # Send verification email
            EmailService.send_email_verification(
                user=user,
                request=request,
                token=token,
                uid=uid
            )
            
            print(f"DEBUG: Verification email resent to {email}")
            
            messages.success(request, 'Verification email sent! Please check your inbox.')
            
        except User.DoesNotExist:
            # Don't reveal if email exists (security)
            messages.success(request, 'If an account exists, a verification email has been sent.')
        
        return redirect('dashboard:resend_verification')
    
    return render(request, 'dashboard/auth/resend_verification.html')

@redirect_authenticated_user
@ensure_csrf_cookie
def register_patient_page(request):
    """Handle patient registration with email verification."""
    if request.method == 'POST':
        email = request.POST.get('email', '').strip().lower()
        password = request.POST.get('password')
        password_confirm = request.POST.get('password_confirm')
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        phone = request.POST.get('phone', '').strip()
        
        # Validation
        errors = []
        
        if not all([email, password, password_confirm, first_name, last_name]):
            errors.append('All required fields must be filled.')
        
        # Validate email format and check for disposable emails
        if email:
            from accounts.email_validator import EmailValidator
            is_valid, error_msg, suggestion = EmailValidator.validate_email(email)
            if not is_valid:
                errors.append(error_msg)
        
        if password != password_confirm:
            errors.append('Passwords do not match.')
        
        if password and len(password) < 8:
            errors.append('Password must be at least 8 characters.')
        
        if email and User.objects.filter(email=email).exists():
            errors.append('An account with this email already exists.')
        
        if errors:
            for error in errors:
                messages.error(request, error)
            return render(request, 'dashboard/auth/register_patient.html', {
                'email': email,
                'first_name': first_name,
                'last_name': last_name,
                'phone': phone,
            })
        
        # Create user (email not verified yet)
        user = User.objects.create_user(
            email=email,
            password=password,
            first_name=first_name,
            last_name=last_name,
            user_type='patient',
            phone=phone,
            is_active=True,
            email_verified=False,  # ✅ Not verified yet
        )
        
        # Create patient profile
        PatientProfile.objects.create(user=user)
        
        # Generate verification token
        token = default_token_generator.make_token(user)
        uid = urlsafe_base64_encode(force_bytes(user.pk))
        
        # Send verification email
        EmailService.send_email_verification(
            user=user,
            request=request,
            token=token,
            uid=uid
        )
        
        # Send welcome email
        EmailService.send_welcome_email(user)
        
        print(f"DEBUG: Verification email sent to {email}")
        
        messages.success(
            request, 
            'Account created successfully! Please check your email to verify your account.'
        )
        
        # ✅ Redirect to verification sent page, NOT dashboard
        return redirect('dashboard:verification_sent')
    
    return render(request, 'dashboard/auth/register_patient.html')


@redirect_authenticated_user
@ensure_csrf_cookie
def register_doctor_page(request):
    """Handle doctor registration with email verification."""
    if request.method == 'POST':
        email = request.POST.get('email', '').strip().lower()
        password = request.POST.get('password')
        password_confirm = request.POST.get('password_confirm')
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        phone = request.POST.get('phone', '').strip()
        license_number = request.POST.get('license_number', '').strip()
        specialization_id = request.POST.get('specialization')
        experience_years = request.POST.get('experience_years', 0)
        consultation_fee = request.POST.get('consultation_fee', 0)
        education = request.POST.get('education', '').strip()
        
        # Validation
        errors = []
        
        if not all([email, password, password_confirm, first_name, last_name, license_number]):
            errors.append('All required fields must be filled.')
        
        # Validate email format and check for disposable emails
        if email:
            from accounts.email_validator import EmailValidator
            is_valid, error_msg, suggestion = EmailValidator.validate_email(email)
            if not is_valid:
                errors.append(error_msg)
        
        if password != password_confirm:
            errors.append('Passwords do not match.')
        
        if password and len(password) < 8:
            errors.append('Password must be at least 8 characters.')
        
        if email and User.objects.filter(email=email).exists():
            errors.append('An account with this email already exists.')
        
        if DoctorProfile.objects.filter(license_number=license_number).exists():
            errors.append('This license number is already registered.')
        
        if errors:
            for error in errors:
                messages.error(request, error)
            
            # Get specializations for the form
            from doctors.models import Specialization
            specializations = Specialization.objects.all()
            
            return render(request, 'dashboard/auth/register_doctor.html', {
                'email': email,
                'first_name': first_name,
                'last_name': last_name,
                'phone': phone,
                'license_number': license_number,
                'experience_years': experience_years,
                'consultation_fee': consultation_fee,
                'education': education,
                'specializations': specializations,
            })
        
        # Create user (email not verified yet)
        user = User.objects.create_user(
            email=email,
            password=password,
            first_name=first_name,
            last_name=last_name,
            user_type='doctor',
            phone=phone,
            is_active=True,
            email_verified=False,  # ✅ Not verified yet
        )
        
        # Get specialization
        from doctors.models import Specialization
        specialization = None
        if specialization_id:
            try:
                specialization = Specialization.objects.get(id=specialization_id)
            except Specialization.DoesNotExist:
                pass
        
        # Create doctor profile (pending verification)
        DoctorProfile.objects.create(
            user=user,
            license_number=license_number,
            specialization=specialization,
            experience_years=experience_years or 0,
            consultation_fee=consultation_fee or 0,
            education=education,
            verification_status='pending',  # Doctor needs admin approval too
        )
        
        # Generate verification token
        token = default_token_generator.make_token(user)
        uid = urlsafe_base64_encode(force_bytes(user.pk))
        
        # Send verification email
        EmailService.send_email_verification(
            user=user,
            request=request,
            token=token,
            uid=uid
        )
        
        # Send welcome email
        EmailService.send_welcome_email(user)
        
        print(f"DEBUG: Verification email sent to {email}")
        
        messages.success(
            request, 
            'Account created successfully! Please check your email to verify your account. '
            'Your doctor profile will be reviewed by our team.'
        )
        
        # ✅ Redirect to verification sent page, NOT dashboard
        return redirect('dashboard:verification_sent')
    
    # GET request - show form with specializations
    from doctors.models import Specialization
    specializations = Specialization.objects.all()
    
    return render(request, 'dashboard/auth/register_doctor.html', {
        'specializations': specializations,
    })


@redirect_authenticated_user
def forgot_password_page(request):
    """Forgot password - Send reset link via email."""
    if request.method == 'POST':
        email = request.POST.get('email', '').strip().lower()
        
        try:
            user = User.objects.get(email=email)
            
            # Generate password reset token
            token = default_token_generator.make_token(user)
            uid = urlsafe_base64_encode(force_bytes(user.pk))
            
            # ✅ Use EmailService
            EmailService.send_password_reset(
                user=user,
                request=request,
                token=token,
                uid=uid
            )
            
            print(f"DEBUG: Password reset email sent to {email}")
            
        except User.DoesNotExist:
            pass
        
        messages.success(
            request, 
            'If an account exists with this email, you will receive password reset instructions shortly.'
        )
        return redirect('dashboard:forgot_password')
    
    return render(request, 'dashboard/auth/forgot_password.html')

@redirect_authenticated_user
def reset_password_page(request, uidb64, token):
    """Reset password using token from email."""
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = User.objects.get(pk=uid)
    except (TypeError, ValueError, OverflowError, User.DoesNotExist):
        user = None
    
    # Validate token
    if user is None or not default_token_generator.check_token(user, token):
        messages.error(request, 'This password reset link is invalid or has expired.')
        return redirect('dashboard:forgot_password')
    
    if request.method == 'POST':
        password1 = request.POST.get('password')
        password2 = request.POST.get('password_confirm')
        
        # Validation
        if not password1 or not password2:
            messages.error(request, 'Please fill in both password fields.')
        elif password1 != password2:
            messages.error(request, 'Passwords do not match.')
        elif len(password1) < 8:
            messages.error(request, 'Password must be at least 8 characters long.')
        else:
            # Set new password
            user.set_password(password1)
            user.save()
            
            messages.success(request, 'Your password has been reset successfully! Please log in.')
            return redirect('dashboard:login')
    
    return render(request, 'dashboard/auth/reset_password.html', {
        'uidb64': uidb64,
        'token': token,
    })


def logout_page(request):
    """Logout and redirect to login page."""
    # Clear all existing messages before logout
    storage = messages.get_messages(request)
    storage.used = True  # Mark all messages as used/consumed
    
    logout(request)
    messages.success(request, 'You have been logged out successfully.')
    return redirect('dashboard:login')

# ==========================
# DASHBOARD VIEWS
# ==========================

@login_required
@doctor_required
def doctor_dashboard(request):
    """Doctor Dashboard with real data."""
    
    user = request.user
    doctor_profile = user.doctor_profile
    today = timezone.now().date()
    
    # =============================================================================
    # STATS CARDS
    # =============================================================================
    
    # Total Appointments (all time)
    total_appointments = Appointment.objects.filter(doctor=doctor_profile).count()
    
    # Today's Appointments
    todays_appointments = Appointment.objects.filter(
        doctor=doctor_profile,
        date=today
    ).exclude(status='cancelled')
    todays_count = todays_appointments.count()
    
    # This Week's Appointments
    week_start = today - timedelta(days=today.weekday())
    week_end = week_start + timedelta(days=6)
    this_week_appointments = Appointment.objects.filter(
        doctor=doctor_profile,
        date__gte=week_start,
        date__lte=week_end
    ).exclude(status='cancelled').count()
    
    # Total Patients (unique)
    total_patients = Appointment.objects.filter(
        doctor=doctor_profile
    ).values('patient').distinct().count()
    
    # Completed Consultations
    completed_consultations = Appointment.objects.filter(
        doctor=doctor_profile,
        status='completed'
    ).count()
    
    # Total Earnings (completed appointments)
    total_earnings = Appointment.objects.filter(
        doctor=doctor_profile,
        status='completed'
    ).count() * float(doctor_profile.consultation_fee)
    
    # This Month's Earnings
    month_start = today.replace(day=1)
    monthly_earnings = Appointment.objects.filter(
        doctor=doctor_profile,
        status='completed',
        date__gte=month_start
    ).count() * float(doctor_profile.consultation_fee)
    
    # Pending Appointments (need confirmation or action)
    pending_appointments = Appointment.objects.filter(
        doctor=doctor_profile,
        status='pending'
    ).count()
    
    # =============================================================================
    # TODAY'S SCHEDULE
    # =============================================================================
    
    todays_schedule = Appointment.objects.filter(
        doctor=doctor_profile,
        date=today
    ).exclude(
        status='cancelled'
    ).select_related(
        'patient'
    ).order_by('start_time')[:10]
    
    # =============================================================================
    # UPCOMING APPOINTMENTS (Next 7 days, excluding today)
    # =============================================================================
    
    upcoming_appointments = Appointment.objects.filter(
        doctor=doctor_profile,
        date__gt=today,
        date__lte=today + timedelta(days=7),
        status__in=['confirmed', 'pending']
    ).select_related(
        'patient'
    ).order_by('date', 'start_time')[:5]
    
    # =============================================================================
    # RECENT PATIENTS
    # =============================================================================
    
    recent_patients = Appointment.objects.filter(
        doctor=doctor_profile,
        status='completed'
    ).select_related(
        'patient', 'patient__patient_profile'
    ).order_by('-date', '-end_time')[:5]
    
    # Get unique patients
    seen_patients = set()
    unique_recent_patients = []
    for apt in recent_patients:
        if apt.patient.id not in seen_patients:
            seen_patients.add(apt.patient.id)
            unique_recent_patients.append(apt)
    
    # =============================================================================
    # PENDING REVIEWS (Completed appointments without consultation notes)
    # =============================================================================
    
    pending_reviews = Appointment.objects.filter(
        doctor=doctor_profile,
        status='completed'
    ).exclude(
        consultation__isnull=False
    ).select_related('patient').order_by('-date')[:5]
    
    # =============================================================================
    # RECENT ACTIVITY
    # =============================================================================
    
    recent_activity = []
    
    # Recent appointments (last 7 days)
    recent_apts = Appointment.objects.filter(
        doctor=doctor_profile,
        created_at__gte=timezone.now() - timedelta(days=7)
    ).select_related('patient').order_by('-created_at')[:10]
    
    for apt in recent_apts:
        if apt.status == 'confirmed':
            activity_type = 'New Appointment'
            icon_class = 'text-success'
        elif apt.status == 'completed':
            activity_type = 'Consultation Completed'
            icon_class = 'text-primary'
        elif apt.status == 'cancelled':
            activity_type = 'Appointment Cancelled'
            icon_class = 'text-danger'
        else:
            activity_type = 'Appointment Updated'
            icon_class = 'text-warning'
        
        recent_activity.append({
            'type': activity_type,
            'description': f'{apt.patient.full_name}',
            'timestamp': apt.updated_at,
            'icon_class': icon_class,
        })
    
    # =============================================================================
    # CONTEXT
    # =============================================================================
    
    context = {
        'user': user,
        'profile': doctor_profile,
        
        # Stats
        'total_appointments': total_appointments,
        'todays_count': todays_count,
        'this_week_appointments': this_week_appointments,
        'total_patients': total_patients,
        'completed_consultations': completed_consultations,
        'total_earnings': total_earnings,
        'monthly_earnings': monthly_earnings,
        'pending_appointments': pending_appointments,
        
        # Lists
        'todays_schedule': todays_schedule,
        'upcoming_appointments': upcoming_appointments,
        'recent_patients': unique_recent_patients,
        'pending_reviews': pending_reviews,
        'recent_activity': recent_activity,
        
        # Helpers
        'today': today,
    }
    
    return render(request, 'dashboard/doctor/dashboard.html', context)

@login_required
@doctor_required
def doctor_change_password(request):
    """Change password for logged-in doctors."""
    if request.method == 'POST':
        current_password = request.POST.get('current_password')
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')
        
        # Validation
        if not request.user.check_password(current_password):
            messages.error(request, 'Current password is incorrect.')
        elif not new_password or not confirm_password:
            messages.error(request, 'Please fill in all password fields.')
        elif new_password != confirm_password:
            messages.error(request, 'New passwords do not match.')
        elif len(new_password) < 8:
            messages.error(request, 'Password must be at least 8 characters long.')
        elif current_password == new_password:
            messages.error(request, 'New password must be different from current password.')
        else:
            # Update password
            request.user.set_password(new_password)
            request.user.save()
            
            # Keep user logged in after password change
            update_session_auth_hash(request, request.user)
            
            messages.success(request, 'Your password has been changed successfully!')
            return redirect('dashboard:doctor_change_password')
    
    return render(request, 'dashboard/doctor/change_password.html')


@login_required
@doctor_required
def doctor_appointments(request):
    """List appointments for logged-in doctor."""
    doctor_profile = request.user.doctor_profile

    appointments = (
        Appointment.objects
        .filter(doctor=doctor_profile)
        .select_related('patient')
        .order_by('-date', '-start_time')
    )

    # Simple search
    q = request.GET.get('q')
    if q:
        appointments = appointments.filter(
            Q(patient__first_name__icontains=q) |
            Q(patient__last_name__icontains=q) |
            Q(patient__email__icontains=q) |
            Q(appointment_number__icontains=q)
        )

    # Filter by status
    status = request.GET.get('status')
    if status:
        appointments = appointments.filter(status=status)

    # Filter by date
    date_filter = request.GET.get('date')
    if date_filter:
        try:
            filter_date = datetime.strptime(date_filter, '%Y-%m-%d').date()
            appointments = appointments.filter(date=filter_date)
        except ValueError:
            pass

    # Get all patients for the new appointment form
    patients = User.objects.filter(user_type='patient', is_active=True).order_by('first_name', 'last_name')

    return render(request, 'dashboard/doctor/appointments.html', {
        'appointments': appointments,
        'patients': patients,
    })


@login_required
@doctor_required
def doctor_appointment_calendar(request):
    """Renders the full calendar view of appointments."""
    
    # Get all patients for the new appointment form
    patients = User.objects.filter(user_type='patient', is_active=True).order_by('first_name', 'last_name')
    
    return render(request, 'dashboard/doctor/appointment_calendar.html', {
        'patients': patients,
    })


@login_required
@doctor_required
def doctor_appointment_detail(request, pk):
    """View appointment details."""
    doctor_profile = request.user.doctor_profile
    appointment = get_object_or_404(Appointment, pk=pk, doctor=doctor_profile)

    try:
        consultation = appointment.consultation
    except Consultation.DoesNotExist:
        consultation = None

    prescriptions = consultation.prescriptions.all() if consultation else Prescription.objects.none()

    context = {
        'appointment': appointment,
        'patient': appointment.patient,
        'consultation': consultation,
        'prescriptions': prescriptions,
    }
    return render(request, 'dashboard/doctor/appointment_detail.html', context)


@login_required
@doctor_required
def doctor_appointment_events(request):
    """JSON feed of appointments for FullCalendar."""
    doctor_profile = request.user.doctor_profile

    start = request.GET.get('start')
    end = request.GET.get('end')

    qs = Appointment.objects.filter(doctor=doctor_profile).exclude(status='cancelled')

    if start:
        qs = qs.filter(date__gte=start[:10])  # FullCalendar sends ISO datetime, we need just date
    if end:
        qs = qs.filter(date__lte=end[:10])

    events = []
    for apt in qs.select_related('patient'):
        start_str = f"{apt.date}T{apt.start_time.strftime('%H:%M:%S')}"
        end_str = f"{apt.date}T{apt.end_time.strftime('%H:%M:%S')}"

        # Color by status
        if apt.status in ['confirmed', 'in_progress']:
            color = '#0d6efd'  # blue
        elif apt.status == 'completed':
            color = '#198754'  # green
        elif apt.status == 'pending':
            color = '#ffc107'  # yellow
        else:
            color = '#6c757d'  # gray

        events.append({
            'id': apt.id,
            'title': apt.patient.full_name,
            'start': start_str,
            'end': end_str,
            'url': reverse('dashboard:doctor_appointment_detail', args=[apt.id]),
            'backgroundColor': color,
            'borderColor': color,
            'extendedProps': {
                'status': apt.status,
                'appointment_number': apt.appointment_number,
                'mode': 'online' if apt.video_room_url else 'in_person',
            },
        })

    return JsonResponse(events, safe=False)

@login_required
@doctor_required
def doctor_confirm_appointment(request, pk):
    """Doctor confirms a pending appointment."""
    doctor_profile = request.user.doctor_profile
    
    try:
        appointment = Appointment.objects.get(
            pk=pk, 
            doctor=doctor_profile,
            status='pending'
        )
    except Appointment.DoesNotExist:
        messages.error(request, 'Appointment not found or already processed.')
        return redirect('dashboard:doctor_appointments')
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'confirm':
            appointment.status = 'confirmed'
            
            # Ensure video room exists for online appointments (if it was created during booking)
            # Video room should already exist from booking, but generate if missing
            if not appointment.video_room_url:
                # Check if this was meant to be an online appointment
                # (We can't determine this from the model, but if video_room_url is missing,
                # it might have failed during booking, so try to generate it now)
                try:
                    appointment.generate_video_room()
                except Exception as e:
                    print(f"Warning: Failed to generate video room: {e}")
                    # Don't fail the confirmation if video room generation fails
            
            appointment.save()
            
            # Send confirmation email to patient
            try:
                from notifications.services import EmailService
                EmailService.send_appointment_confirmed(appointment)
            except Exception as e:
                print(f"Failed to send confirmation email: {e}")
            
            messages.success(request, f'Appointment {appointment.appointment_number} confirmed!')
            
        elif action == 'reject':
            rejection_reason = request.POST.get('rejection_reason', 'No reason provided')
            appointment.status = 'cancelled'
            appointment.cancellation_reason = rejection_reason
            appointment.cancelled_by = request.user
            appointment.cancelled_at = timezone.now()
            appointment.save()
            
            # Send rejection email to patient
            try:
                from notifications.services import EmailService
                EmailService.send_appointment_rejected(appointment, rejection_reason)
            except Exception as e:
                print(f"Failed to send rejection email: {e}")
            
            messages.info(request, f'Appointment {appointment.appointment_number} has been declined.')
        
        return redirect('dashboard:doctor_appointments')
    
    # GET request - show confirmation page
    context = {
        'appointment': appointment,
    }
    return render(request, 'dashboard/doctor/confirm_appointment.html', context)


@login_required
@doctor_required
def doctor_pending_appointments(request):
    """List all pending appointments for doctor to review."""
    doctor_profile = request.user.doctor_profile
    
    pending_appointments = Appointment.objects.filter(
        doctor=doctor_profile,
        status='pending'
    ).select_related('patient').order_by('date', 'start_time')
    
    context = {
        'appointments': pending_appointments,
        'pending_count': pending_appointments.count(),
    }
    return render(request, 'dashboard/doctor/pending_appointments.html', context)

@login_required
@doctor_required
def doctor_create_appointment(request):
    """Create a new appointment."""
    if request.method == 'POST':
        doctor_profile = request.user.doctor_profile
        
        patient_id = request.POST.get('patient')
        date_str = request.POST.get('date')
        start_time_str = request.POST.get('start_time')
        duration = int(request.POST.get('duration', 30))
        reason = request.POST.get('reason', '')
        status = request.POST.get('status', 'confirmed')
        appointment_type = request.POST.get('appointment_type', 'online')
        
        # Validate patient
        try:
            patient = User.objects.get(id=patient_id, user_type='patient')
        except User.DoesNotExist:
            messages.error(request, 'Invalid patient selected.')
            return redirect('dashboard:doctor_appointments')
        
        # Parse date and time
        try:
            date = datetime.strptime(date_str, '%Y-%m-%d').date()
            start_time = datetime.strptime(start_time_str, '%H:%M').time()
            
            # Calculate end time
            start_datetime = datetime.combine(date, start_time)
            end_datetime = start_datetime + timedelta(minutes=duration)
            end_time = end_datetime.time()
        except ValueError:
            messages.error(request, 'Invalid date or time format.')
            return redirect('dashboard:doctor_appointments')
        
        # Check for conflicting appointments
        conflicting = Appointment.objects.filter(
            doctor=doctor_profile,
            date=date,
            status__in=['pending', 'confirmed', 'in_progress']
        ).filter(
            Q(start_time__lt=end_time, end_time__gt=start_time)
        ).exists()
        
        if conflicting:
            messages.error(request, 'This time slot conflicts with an existing appointment.')
            return redirect('dashboard:doctor_appointments')
        
        # Create the appointment
        appointment = Appointment.objects.create(
            patient=patient,
            doctor=doctor_profile,
            date=date,
            start_time=start_time,
            end_time=end_time,
            reason=reason,
            status=status,
        )
        
        # Generate video room if online consultation
        if appointment_type == 'online':
            appointment.generate_video_room()
            appointment.save()
        
        messages.success(request, f'Appointment {appointment.appointment_number} created successfully!')
        return redirect('dashboard:doctor_appointments')
    
    return redirect('dashboard:doctor_appointments')


@login_required
@doctor_required
def doctor_cancel_appointment(request, pk):
    """Cancel an appointment."""
    if request.method == 'POST':
        doctor_profile = request.user.doctor_profile
        appointment = get_object_or_404(Appointment, pk=pk, doctor=doctor_profile)
        
        if not appointment.can_cancel:
            messages.error(request, 'This appointment cannot be cancelled.')
            return redirect('dashboard:doctor_appointment_detail', pk=pk)
        
        cancellation_reason = request.POST.get('cancellation_reason', '')
        
        appointment.status = 'cancelled'
        appointment.cancellation_reason = cancellation_reason
        appointment.cancelled_by = request.user
        appointment.cancelled_at = timezone.now()
        appointment.save()
        
        messages.success(request, f'Appointment {appointment.appointment_number} has been cancelled.')
        return redirect('dashboard:doctor_appointments')
    
    return redirect('dashboard:doctor_appointments')

@doctor_required
@login_required
def doctor_patients(request):
    if request.user.user_type != 'doctor':
        return redirect('dashboard')

    # CORRECT WAY — using the real related_name: patient_appointments
    queryset = User.objects.filter(
        user_type='patient',
        patient_appointments__doctor=request.user.doctor_profile  # ← THIS IS CORRECT
    ).distinct().select_related('patient_profile')

    # Search filter
    q = request.GET.get('q')
    gender = request.GET.get('gender')

    if q:
        queryset = queryset.filter(
            models.Q(first_name__icontains=q) |
            models.Q(last_name__icontains=q) |
            models.Q(email__icontains=q)
        )
    if gender:
        queryset = queryset.filter(gender=gender)

    patients = queryset

    context = {
        'patients': patients
    }
    return render(request, 'dashboard/doctor/patients.html', context)

@login_required
@doctor_required
def doctor_patient_detail(request, pk):
    patient = get_object_or_404(User, id=pk, user_type='patient')
    
    # SECURITY CHECK - FIXED!
    has_access = Appointment.objects.filter(
        doctor=request.user.doctor_profile,  # ← THIS IS CORRECT
        patient=patient
    ).exists()

    if not has_access and not request.user.is_superuser:
        messages.error(request, "You are not authorized to view this patient.")
        return redirect('dashboard:doctor_patients')

    # Get appointments with this patient
    appointments = Appointment.objects.filter(
        doctor=request.user.doctor_profile,  # ← ALSO FIX HERE
        patient=patient
    ).order_by('-date', '-start_time')

    context = {
        'patient': patient,
        'appointments': appointments,
        'prescriptions': Prescription.objects.filter(consultation__appointment__patient=patient, consultation__appointment__doctor=request.user.doctor_profile).order_by('-created_at'),
    }
    return render(request, 'dashboard/doctor/patient_detail.html', context)


@login_required
@doctor_required
def doctor_patient_records(request, pk):
    """Doctor views patient's medical records"""
    patient = get_object_or_404(User, pk=pk, user_type='patient')
    
    # Security: Check doctor has treated this patient
    has_access = Appointment.objects.filter(
        doctor=request.user.doctor_profile,
        patient=patient
    ).exists()
    
    if not has_access and not request.user.is_superuser:
        messages.error(request, "You don't have access to this patient's records.")
        return redirect('dashboard:doctor_patients')
    
    # Get all records
    health_profile = HealthProfile.objects.filter(patient=patient).first()
    medical_history = MedicalHistory.objects.filter(patient=patient).order_by('-event_date')
    documents = MedicalDocument.objects.filter(patient=patient).order_by('-uploaded_at')
    
    context = {
        'patient': patient,
        'health_profile': health_profile,
        'medical_history': medical_history,
        'documents': documents,
    }
    return render(request, 'dashboard/doctor/patient_records.html', context)


@login_required
@doctor_required
def doctor_patient_document_view(request, patient_id, doc_id):
    """Doctor views a patient's document with signed URL"""
    patient = get_object_or_404(User, pk=patient_id, user_type='patient')
    
    # Security check
    has_access = Appointment.objects.filter(
        doctor=request.user.doctor_profile,
        patient=patient
    ).exists()
    
    if not has_access and not request.user.is_superuser:
        messages.error(request, "Access denied.")
        return redirect('dashboard:doctor_patients')
    
    document = get_object_or_404(MedicalDocument, pk=doc_id, patient=patient)
    
    # Generate signed URL
    signed_url = generate_signed_url(document.file.name)
    
    if signed_url:
        return HttpResponseRedirect(signed_url)
    else:
        messages.error(request, 'Could not generate download link.')
        return redirect('dashboard:doctor_patient_records', pk=patient_id)

@login_required
@doctor_required
def doctor_schedule(request):
    return render(request, 'dashboard/doctor/schedule.html')

@login_required
@doctor_required
def doctor_prescriptions(request):
    """
    List prescriptions written by the logged-in doctor.
    Optionally can be filtered by patient in the future.
    """
    doctor_profile = request.user.doctor_profile

    prescriptions = Prescription.objects.filter(
        consultation__appointment__doctor=doctor_profile
    ).select_related(
        'consultation__appointment__patient'
    ).order_by('-created_at')

    # Search functionality
    search_query = request.GET.get('q', '').strip()
    if search_query:
        prescriptions = prescriptions.filter(
            Q(prescription_number__icontains=search_query) |
            Q(consultation__appointment__patient__first_name__icontains=search_query) |
            Q(consultation__appointment__patient__last_name__icontains=search_query) |
            Q(consultation__appointment__patient__email__icontains=search_query)
        )

    # Ensure prescription_number exists for all prescriptions
    for pres in prescriptions:
        if not pres.prescription_number:
            pres.save()  # This will trigger auto-generation

    context = {
        'prescriptions': prescriptions,
    }
    return render(request, 'dashboard/doctor/prescriptions.html', context)

@login_required
@doctor_required
def doctor_prescription_create(request, patient_id):
    patient = get_object_or_404(User, id=patient_id, user_type='patient')
    
    # Security: doctor can only prescribe to patients they have seen
    if not Appointment.objects.filter(
        doctor=request.user.doctor_profile,
        patient=patient
    ).exists():
        messages.error(request, "You can only prescribe to your own patients.")
        return redirect('dashboard:doctor_patients')

    if request.method == 'POST':
        diagnosis = request.POST.get('diagnosis', '').strip()
        medicines_text = request.POST.get('medicines', '').strip()
        instructions = request.POST.get('instructions', '').strip()

        # Find the most recent completed appointment between this doctor and patient
        appointment = (
            Appointment.objects
            .filter(
                doctor=request.user.doctor_profile,
                patient=patient,
                status='completed',
            )
            .order_by('-date', '-start_time')
            .first()
        )

        if not appointment:
            messages.error(
                request,
                "You can only issue prescriptions for completed consultations."
            )
            return redirect('dashboard:doctor_patient_detail', pk=patient.id)

        # Get or create consultation for that appointment
        consultation, _ = Consultation.objects.get_or_create(appointment=appointment)

        # Create structured prescription
        prescription = Prescription.objects.create(
            consultation=consultation,
            diagnosis=diagnosis,
            notes=instructions,
        )

        # Store the free-text medicines as a single item so it shows up in UI
        if medicines_text:
            from consultations.models import PrescriptionItem

            PrescriptionItem.objects.create(
                prescription=prescription,
                medicine_name="Custom regimen",
                dosage="",
                frequency="other",
                duration="other",
                instructions=medicines_text,
            )

        messages.success(request, "Prescription created successfully!")
        return redirect('dashboard:doctor_prescription_detail', pk=prescription.pk)

    context = {
        'patient': patient,
        'page_title': f"New Prescription for {patient.full_name}"
    }
    return render(request, 'dashboard/doctor/prescription_create.html', context)

@login_required
@doctor_required
def doctor_prescriptions_export(request, format='pdf'):
    """Export prescriptions as PDF or Excel"""
    from django.http import HttpResponse
    from datetime import datetime
    
    doctor_profile = request.user.doctor_profile
    
    prescriptions = Prescription.objects.filter(
        consultation__appointment__doctor=doctor_profile
    ).select_related(
        'consultation__appointment__patient'
    ).order_by('-created_at')
    
    # Search filter if provided
    search_query = request.GET.get('q', '').strip()
    if search_query:
        prescriptions = prescriptions.filter(
            Q(prescription_number__icontains=search_query) |
            Q(consultation__appointment__patient__first_name__icontains=search_query) |
            Q(consultation__appointment__patient__last_name__icontains=search_query) |
            Q(consultation__appointment__patient__email__icontains=search_query)
    )
    
    # Ensure prescription_number exists
    for pres in prescriptions:
        if not pres.prescription_number:
            pres.save()
    
    if format == 'excel':
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            
            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Prescriptions"
            
            # Headers
            headers = ['Prescription ID', 'Patient Name', 'Patient Email', 'Date Created', 'Issued Date', 'Valid Until']
            ws.append(headers)
            
            # Style headers
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Add data
            for pres in prescriptions:
                patient = pres.consultation.appointment.patient
                prescription_id = pres.prescription_number if pres.prescription_number else f"RX-{pres.id:06d}"
                ws.append([
                    prescription_id,
                    patient.full_name,
                    patient.email,
                    pres.created_at.strftime('%Y-%m-%d %H:%M') if pres.created_at else '',
                    pres.issued_date.strftime('%Y-%m-%d') if pres.issued_date else '',
                    pres.valid_until.strftime('%Y-%m-%d') if pres.valid_until else 'N/A',
                ])
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Create response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f"prescriptions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            wb.save(response)
            return response
        except ImportError:
            messages.error(request, "Excel export requires openpyxl. Please install it: pip install openpyxl")
            return redirect('dashboard:doctor_prescriptions')
    
    elif format == 'pdf':
        # For PDF, create HTML that can be printed to PDF
        from django.template.loader import render_to_string
        
        context = {
            'prescriptions': prescriptions,
            'doctor': doctor_profile,
            'export_date': datetime.now(),
        }
        
        html_content = render_to_string('dashboard/doctor/prescriptions_export_pdf.html', context)
        
        response = HttpResponse(html_content, content_type='text/html')
        response['Content-Disposition'] = f'attachment; filename="prescriptions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"'
        return response
    
    return redirect('dashboard:doctor_prescriptions')

@login_required
@doctor_required
def doctor_prescription_detail(request, pk):
    """
    Detailed view of a prescription for the doctor.
    """
    doctor_profile = request.user.doctor_profile

    prescription = get_object_or_404(
        Prescription.objects.select_related(
            'consultation__appointment__doctor__user',
            'consultation__appointment__patient__patient_profile',
        ),
        pk=pk,
        consultation__appointment__doctor=doctor_profile,
    )

    items = prescription.items.all()
    appointment = prescription.consultation.appointment
    patient = appointment.patient

    context = {
        'prescription': prescription,
        'items': items,
        'appointment': appointment,
        'patient': patient,
        'doctor': appointment.doctor,
    }
    return render(request, 'dashboard/doctor/prescription_detail.html', context)

@login_required
@doctor_required
def doctor_profile(request):
    """Handle Doctor Profile View & Update"""
    profile, created = DoctorProfile.objects.get_or_create(user=request.user)

    if request.method == 'POST':
        
        # ✅ Handle profile picture SEPARATELY (direct model save)
        if 'profile_picture' in request.FILES:
            pic = request.FILES['profile_picture']
            print(f"DEBUG: Uploading file: {pic.name}, Size: {pic.size}")
            
            # Delete old picture from storage (optional but recommended)
            if request.user.profile_picture:
                try:
                    request.user.profile_picture.delete(save=False)
                except Exception as e:
                    print(f"DEBUG: Could not delete old picture: {e}")
            
            # Save new picture directly to user model (uses S3 storage)
            request.user.profile_picture = pic
            request.user.save(update_fields=['profile_picture'])
            print(f"DEBUG: File saved to: {request.user.profile_picture.url}")
        
            # ✅ Debug: Check what URL is being generated
            # if request.user.profile_picture:
            #     print(f"DEBUG: Profile picture name: {request.user.profile_picture.name}")
            #     print(f"DEBUG: Profile picture URL: {request.user.profile_picture.url}")

        # Handle other form data with serializer
        data = request.POST.dict()
        
        user_data = {
            'first_name': data.get('first_name'),
            'last_name': data.get('last_name'),
            'phone': data.get('phone'),
        }
        # ❌ REMOVE: user_data['profile_picture'] = pic (no longer needed)
        
        data['user'] = user_data
        
        serializer = DoctorProfileSerializer(instance=profile, data=data, partial=True)
        
        if serializer.is_valid():
            serializer.save()
            messages.success(request, 'Profile updated successfully!')
            return redirect('dashboard:doctor_profile')
        else:
            for field, errors in serializer.errors.items():
                if isinstance(errors, dict):
                    for subfield, sub_errors in errors.items():
                        messages.error(request, f"{subfield}: {sub_errors[0]}")
                else:
                    messages.error(request, f"{field}: {errors[0]}")

    return render(request, 'dashboard/doctor/profile.html', {
        'user': request.user, 
        'profile': profile
    })

@login_required
@doctor_required
def doctor_notifications(request):
    return render(request, 'dashboard/doctor/notifications.html')


# ==========================
# PATIENT VIEWS
# ==========================

@login_required
@patient_required
def patient_dashboard(request):
    """Patient Dashboard with real data."""
    
    user = request.user
    patient_profile = user.patient_profile
    today = timezone.now().date()
    
    # =============================================================================
    # STATS CARDS
    # =============================================================================
    
    # Total Appointments
    total_appointments = Appointment.objects.filter(patient=user).count()
    
    # Completed Consultations
    completed_consultations = Appointment.objects.filter(
        patient=user,
        status='completed'
    ).count()
    
    # Upcoming Appointments
    upcoming_count = Appointment.objects.filter(
        patient=user,
        date__gte=today,
        status__in=['confirmed', 'pending']
    ).count()
    
    # =============================================================================
    # UPCOMING APPOINTMENTS
    # =============================================================================
    
    upcoming_appointments = Appointment.objects.filter(
        patient=user,
        date__gte=today,
        status__in=['confirmed', 'pending']
    ).select_related(
        'doctor', 'doctor__user', 'doctor__specialization'
    ).order_by('date', 'start_time')[:5]
    
    # =============================================================================
    # MY DOCTORS (Doctors user has consulted with)
    # =============================================================================
    
    my_doctors_apts = Appointment.objects.filter(
        patient=user,
        status='completed'
    ).select_related(
        'doctor', 'doctor__user', 'doctor__specialization'
    ).order_by('-date')
    
    # Get unique doctors
    seen_doctors = set()
    my_doctors = []
    for apt in my_doctors_apts:
        if apt.doctor.id not in seen_doctors:
            seen_doctors.add(apt.doctor.id)
            # Count bookings with this doctor
            booking_count = Appointment.objects.filter(
                patient=user,
                doctor=apt.doctor
            ).count()
            my_doctors.append({
                'doctor': apt.doctor,
                'booking_count': booking_count,
                'last_visit': apt.date,
            })
        if len(my_doctors) >= 5:
            break
    
    # =============================================================================
    # RECENT PRESCRIPTIONS
    # =============================================================================
    
    recent_prescriptions = Prescription.objects.filter(
        consultation__appointment__patient=user
    ).select_related(
        'consultation__appointment__doctor__user',
        'consultation__appointment__doctor__specialization'
    ).order_by('-created_at')[:5]
    
    # =============================================================================
    # RECENT APPOINTMENTS
    # =============================================================================
    
    recent_appointments = Appointment.objects.filter(
        patient=user
    ).select_related(
        'doctor', 'doctor__user', 'doctor__specialization'
    ).order_by('-date', '-start_time')[:10]
    
    # =============================================================================
    # RECENT ACTIVITY
    # =============================================================================
    
    recent_activity = []
    
    recent_apts = Appointment.objects.filter(
        patient=user,
        created_at__gte=timezone.now() - timedelta(days=30)
    ).select_related('doctor__user').order_by('-created_at')[:10]
    
    for apt in recent_apts:
        if apt.status == 'confirmed':
            activity_type = 'Appointment Confirmed'
            icon_class = 'text-success'
        elif apt.status == 'completed':
            activity_type = 'Consultation Completed'
            icon_class = 'text-primary'
        elif apt.status == 'cancelled':
            activity_type = 'Appointment Cancelled'
            icon_class = 'text-danger'
        else:
            activity_type = 'Appointment Booked'
            icon_class = 'text-info'
        
        recent_activity.append({
            'type': activity_type,
            'description': f'Dr. {apt.doctor.user.last_name}',
            'timestamp': apt.updated_at,
            'icon_class': icon_class,
        })
    
    # =============================================================================
    # VITALS (from patient profile)
    # =============================================================================
    
    vitals = {
        'weight': patient_profile.weight_kg,
        'height': patient_profile.height_cm,
        'blood_type': patient_profile.blood_type,
    }
    
    # =============================================================================
    # CONTEXT
    # =============================================================================
    
    context = {
        'user': user,
        'profile': patient_profile,
        
        # Stats
        'total_appointments': total_appointments,
        'completed_consultations': completed_consultations,
        'upcoming_count': upcoming_count,
        
        # Lists
        'upcoming_appointments': upcoming_appointments,
        'my_doctors': my_doctors,
        'recent_prescriptions': recent_prescriptions,
        'recent_appointments': recent_appointments,
        'recent_activity': recent_activity,
        
        # Vitals
        'vitals': vitals,
        
        # Helpers
        'today': today,
        'doctors': get_verified_doctors(),
    }
    
    return render(request, 'dashboard/patient/dashboard.html', context)

@login_required
@patient_required
def patient_change_password(request):
    """Change password for logged-in patients."""
    if request.method == 'POST':
        current_password = request.POST.get('current_password')
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')
        
        # Validation
        if not request.user.check_password(current_password):
            messages.error(request, 'Current password is incorrect.')
        elif not new_password or not confirm_password:
            messages.error(request, 'Please fill in all password fields.')
        elif new_password != confirm_password:
            messages.error(request, 'New passwords do not match.')
        elif len(new_password) < 8:
            messages.error(request, 'Password must be at least 8 characters long.')
        elif current_password == new_password:
            messages.error(request, 'New password must be different from current password.')
        else:
            # Update password
            request.user.set_password(new_password)
            request.user.save()
            
            # Keep user logged in after password change
            update_session_auth_hash(request, request.user)
            
            messages.success(request, 'Your password has been changed successfully!')
            return redirect('dashboard:patient_change_password')
    
    return render(request, 'dashboard/patient/change_password.html')


@login_required
@patient_required
def patient_appointments(request):
    """List appointments for logged-in patient."""
    user = request.user

    appointments = (
        Appointment.objects
        .filter(patient=user)
        .select_related('doctor', 'doctor__user', 'doctor__specialization')
        .order_by('-date', '-start_time')
    )

    # Simple search
    q = request.GET.get('q')
    if q:
        appointments = appointments.filter(
            Q(doctor__user__first_name__icontains=q) |
            Q(doctor__user__last_name__icontains=q) |
            Q(appointment_number__icontains=q)
        )

    # Filter by status
    status = request.GET.get('status')
    if status:
        appointments = appointments.filter(status=status)

    

    return render(request, 'dashboard/patient/appointments.html', {
        'appointments': appointments,
        'doctors': get_verified_doctors(),
    })


@login_required
@patient_required
def patient_appointment_calendar(request):
    """Renders the full calendar view of appointments for patient."""
    
    
    return render(request, 'dashboard/patient/appointment_calendar.html', {
        'doctors': get_verified_doctors(),
    })


@login_required
@patient_required
def patient_appointment_detail(request, pk):
    """View appointment details for patient."""
    user = request.user
    appointment = get_object_or_404(Appointment, pk=pk, patient=user)

    try:
        consultation = appointment.consultation
    except Consultation.DoesNotExist:
        consultation = None

    prescriptions = consultation.prescriptions.all() if consultation else Prescription.objects.none()

    context = {
        'appointment': appointment,
        'doctor': appointment.doctor,
        'consultation': consultation,
        'prescriptions': prescriptions,
        'doctors': get_verified_doctors(),
    }
    return render(request, 'dashboard/patient/appointment_detail.html', context)


@login_required
@patient_required
def patient_appointment_events(request):
    """JSON feed of appointments for FullCalendar (patient view)."""
    user = request.user

    start = request.GET.get('start')
    end = request.GET.get('end')

    qs = Appointment.objects.filter(patient=user).exclude(status='cancelled')

    if start:
        qs = qs.filter(date__gte=start[:10])
    if end:
        qs = qs.filter(date__lte=end[:10])

    events = []
    for apt in qs.select_related('doctor', 'doctor__user', 'doctor__specialization'):
        start_str = f"{apt.date}T{apt.start_time.strftime('%H:%M:%S')}"
        end_str = f"{apt.date}T{apt.end_time.strftime('%H:%M:%S')}"

        # Color by status
        if apt.status in ['confirmed', 'in_progress']:
            color = '#0d6efd'  # blue
        elif apt.status == 'completed':
            color = '#198754'  # green
        elif apt.status == 'pending':
            color = '#ffc107'  # yellow
        else:
            color = '#6c757d'  # gray

        events.append({
            'id': apt.id,
            'title': f"Dr. {apt.doctor.user.last_name}",
            'start': start_str,
            'end': end_str,
            'url': reverse('dashboard:patient_appointment_detail', args=[apt.id]),
            'backgroundColor': color,
            'borderColor': color,
            'extendedProps': {
                'status': apt.status,
                'appointment_number': apt.appointment_number,
                'specialization': apt.doctor.specialization.name if apt.doctor.specialization else 'General',
                'mode': 'online' if apt.video_room_url else 'in_person',
            },
        })

    return JsonResponse(events, safe=False)


@login_required
@patient_required
def patient_create_appointment(request):
    """Create a new appointment (patient booking)."""
    if request.method == 'POST':
        user = request.user
        
        doctor_id = request.POST.get('doctor')
        date_str = request.POST.get('slot_date')
        start_time_str = request.POST.get('slot_start')
        duration = request.POST.get('duration', 30)
        reason = request.POST.get('reason', '')
        symptoms = request.POST.get('symptoms', '')
        appointment_type = request.POST.get('appointment_type', 'online')
        
        # ========================================
        # VALIDATE REQUIRED FIELDS
        # ========================================
        if not doctor_id:
            messages.error(request, 'Please select a doctor.')
            return redirect('dashboard:patient_create_appointment')
        
        if not date_str:
            messages.error(request, 'Please select a date.')
            return redirect('dashboard:patient_create_appointment')
        
        if not start_time_str:
            messages.error(request, 'Please select a time slot.')
            return redirect('dashboard:patient_create_appointment')
        
        # Validate duration is a number
        try:
            duration = int(duration)
        except (ValueError, TypeError):
            duration = 30  # Default to 30 minutes
        
        # ========================================
        # VALIDATE DOCTOR
        # ========================================
        try:
            doctor = DoctorProfile.objects.get(id=doctor_id, verification_status='verified')
        except DoctorProfile.DoesNotExist:
            messages.error(request, 'Invalid doctor selected.')
            return redirect('dashboard:patient_appointments')
        
        # ========================================
        # PARSE DATE AND TIME
        # ========================================
        try:
            date = datetime.strptime(date_str, '%Y-%m-%d').date()
            start_time = datetime.strptime(start_time_str, '%H:%M').time()
            
            # Calculate end time
            start_datetime = datetime.combine(date, start_time)
            end_datetime = start_datetime + timedelta(minutes=duration)
            end_time = end_datetime.time()
        except ValueError as e:
            messages.error(request, f'Invalid date or time format. Please use the date picker.')
            return redirect('dashboard:patient_create_appointment')
        
        # ========================================
        # VALIDATE DATE IS NOT IN PAST
        # ========================================
        if date < timezone.now().date():
            messages.error(request, 'Cannot book appointments in the past.')
            return redirect('dashboard:patient_create_appointment')
        
        # If booking for today, check time hasn't passed
        if date == timezone.now().date() and start_time < timezone.now().time():
            messages.error(request, 'Cannot book appointments for a time that has already passed.')
            return redirect('dashboard:patient_create_appointment')
        
        # ========================================
        # CHECK FOR CONFLICTS
        # ========================================
        # Check for conflicting appointments (for the doctor)
        conflicting = Appointment.objects.filter(
            doctor=doctor,
            date=date,
            status__in=['pending', 'confirmed', 'in_progress']
        ).filter(
            Q(start_time__lt=end_time, end_time__gt=start_time)
        ).exists()
        
        if conflicting:
            messages.error(request, 'This time slot is not available. Please choose another time.')
            return redirect('dashboard:patient_create_appointment')
        
        # Check if patient already has an appointment at this time
        patient_conflicting = Appointment.objects.filter(
            patient=user,
            date=date,
            status__in=['pending', 'confirmed', 'in_progress']
        ).filter(
            Q(start_time__lt=end_time, end_time__gt=start_time)
        ).exists()
        
        if patient_conflicting:
            messages.error(request, 'You already have an appointment at this time.')
            return redirect('dashboard:patient_create_appointment')
        
        # ========================================
        # CREATE APPOINTMENT
        # ========================================
        try:
            appointment = Appointment.objects.create(
                patient=user,
                doctor=doctor,
                date=date,
                start_time=start_time,
                end_time=end_time,
                reason=reason,
                symptoms=symptoms,
                status='pending',
            )
            
            # Generate video room if online consultation
            if appointment_type == 'online':
                try:
                    appointment.generate_video_room()
                    appointment.save()
                except Exception as e:
                    # Log error but don't fail the booking
                    # Video room can be generated later when joining
                    print(f"Warning: Failed to generate video room: {e}")
            
            messages.success(
                request, 
                f'Appointment {appointment.appointment_number} booked successfully! Awaiting confirmation.'
            )
            return redirect('dashboard:patient_appointments')
            
        except Exception as e:
            messages.error(request, f'Failed to create appointment. Please try again.')
            return redirect('dashboard:patient_create_appointment')
    
    # ========================================
    # GET REQUEST - SHOW FORM
    # ========================================
    # If this is a GET request, you should render a form
    # Add this if you have a form template:
    doctors = DoctorProfile.objects.filter(verification_status='verified')
    context = {
        'doctors': doctors,
    }
    return render(request, 'dashboard/patient/create_appointment.html', context)

@login_required
@patient_required
def patient_cancel_appointment(request, pk):
    """Cancel an appointment (patient)."""
    if request.method == 'POST':
        user = request.user
        appointment = get_object_or_404(Appointment, pk=pk, patient=user)
        
        if not appointment.can_cancel:
            messages.error(request, 'This appointment cannot be cancelled.')
            return redirect('dashboard:patient_appointment_detail', pk=pk)
        
        cancellation_reason = request.POST.get('cancellation_reason', '')
        
        appointment.status = 'cancelled'
        appointment.cancellation_reason = cancellation_reason
        appointment.cancelled_by = user
        appointment.cancelled_at = timezone.now()
        appointment.save()
        
        messages.success(request, f'Appointment {appointment.appointment_number} has been cancelled.')
        return redirect('dashboard:patient_appointments')
    
    return redirect('dashboard:patient_appointments')

@login_required
@patient_required
def patient_doctors(request):
    """Patient sees all verified doctors with filters & search"""
    
    doctors = DoctorProfile.objects.filter(
        verification_status='verified'
    ).select_related('user', 'specialization').order_by('-id')

    # Search
    q = request.GET.get('q', '').strip()
    if q:
        doctors = doctors.filter(
            Q(user__first_name__icontains=q) |
            Q(user__last_name__icontains=q) |
            Q(user__email__icontains=q) |
            Q(specialization__name__icontains=q)
        )

    # Filters
    specialization_id = request.GET.get('specialization')
    if specialization_id:
        doctors = doctors.filter(specialization_id=specialization_id)

    min_experience = request.GET.get('min_experience')
    if min_experience:
        doctors = doctors.filter(experience_years__gte=min_experience)

    max_fee = request.GET.get('max_fee')
    if max_fee:
        doctors = doctors.filter(consultation_fee__lte=max_fee)

    # Sorting
    sort = request.GET.get('sort', 'recent')
    sort_mapping = {
        'name_asc': 'user__first_name',
        'name_desc': '-user__first_name',
        'experience': '-experience_years',
        'rating': '-average_rating',
        'fee_low': 'consultation_fee',
        'fee_high': '-consultation_fee',
    }
    if sort in sort_mapping:
        doctors = doctors.order_by(sort_mapping[sort])

    # Calculate last visit date for each doctor
    from appointments.models import Appointment
    for doctor in doctors:
        last_appointment = Appointment.objects.filter(
            doctor=doctor,
            patient=request.user,
            status='completed'
        ).order_by('-date', '-start_time').first()
        
        if last_appointment:
            doctor.last_visit_date = last_appointment.date
        else:
            doctor.last_visit_date = None

    context = {
        'doctors': doctors,
        'specializations': Specialization.objects.filter(is_active=True),
        'current_q': q,
        'current_specialization': specialization_id,
        'current_min_experience': min_experience,
        'current_max_fee': max_fee,
        'current_sort': sort,
    }
    return render(request, 'dashboard/patient/doctors.html', context)

@login_required
@patient_required
def patient_doctor_detail(request, pk):
    """Single doctor profile for patient"""
    doctor = get_object_or_404(
        DoctorProfile.objects.select_related('user', 'specialization'),
        pk=pk,
        verification_status='verified'
    )

    # Get doctor's weekly availability
    availabilities = Availability.objects.filter(
        doctor=doctor, is_active=True
    ).order_by('day_of_week')

    # Generate next 7 days slots (available only)
    today = date.today()
    next_7_days = [today + timedelta(days=i) for i in range(7)]

    slots_by_date = {}
    for single_date in next_7_days:
        day_of_week = single_date.weekday()  # 0 = Monday
        slots = TimeSlot.objects.filter(
            doctor=doctor,
            date=single_date,
            status='available'
        ).order_by('start_time')

        if slots.exists():
            slots_by_date[single_date] = slots

    context = {
        'doctor': doctor,
        'availabilities': availabilities,
        'slots_by_date': slots_by_date,
        'next_7_days': next_7_days,
        'today': today,
    }
    return render(request, 'dashboard/patient/doctor_detail.html', context)

@login_required
@patient_required
def patient_prescriptions(request):
    prescriptions = Prescription.objects.filter(
        consultation__appointment__patient=request.user
    ).select_related(
        'consultation__appointment__doctor__user',
        'consultation__appointment__doctor__specialization'
    ).order_by('-created_at')

    context = {'prescriptions': prescriptions}
    return render(request, 'dashboard/patient/prescriptions.html', context)

@login_required
@patient_required
def patient_prescription_detail(request, pk):
    prescription = get_object_or_404(
        Prescription.objects.select_related(
            'consultation__appointment__doctor__user',
            'consultation__appointment__doctor__specialization'
        ),
        pk=pk,
        consultation__appointment__patient=request.user
    )

    items = prescription.items.all()

    context = {
        'prescription': prescription,
        'items': items,
        'doctor': prescription.consultation.appointment.doctor,
        'appointment': prescription.consultation.appointment,
    }
    return render(request, 'dashboard/patient/prescription_detail.html', context)

@login_required
@patient_required
def patient_prescription_download(request, pk):
    """Download prescription as PDF - redirects to detail page with print option"""
    # For now, redirect to detail page where user can use browser's print-to-PDF
    # In production, you can integrate a PDF library like weasyprint or reportlab
    return redirect('dashboard:patient_prescription_detail', pk=pk)

@login_required
@patient_required
def patient_prescriptions_export(request, format='pdf'):
    """Export patient prescriptions as PDF or Excel"""
    from django.http import HttpResponse
    from datetime import datetime
    
    prescriptions = Prescription.objects.filter(
        consultation__appointment__patient=request.user
    ).select_related(
        'consultation__appointment__doctor__user',
        'consultation__appointment__doctor__specialization'
    ).order_by('-created_at')
    
    # Ensure prescription_number exists
    for pres in prescriptions:
        if not pres.prescription_number:
            pres.save()
    
    if format == 'excel':
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            
            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "My Prescriptions"
            
            # Headers
            headers = ['Prescription ID', 'Doctor Name', 'Specialization', 'Date Created', 'Issued Date', 'Valid Until']
            ws.append(headers)
            
            # Style headers
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Add data
            for pres in prescriptions:
                doctor = pres.consultation.appointment.doctor
                prescription_id = pres.prescription_number if pres.prescription_number else f"RX-{pres.id:06d}"
                ws.append([
                    prescription_id,
                    f"Dr. {doctor.user.full_name}",
                    doctor.specialization.name,
                    pres.created_at.strftime('%Y-%m-%d %H:%M') if pres.created_at else '',
                    pres.issued_date.strftime('%Y-%m-%d') if pres.issued_date else '',
                    pres.valid_until.strftime('%Y-%m-%d') if pres.valid_until else 'N/A',
                ])
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Create response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f"my_prescriptions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            wb.save(response)
            return response
        except ImportError:
            messages.error(request, "Excel export requires openpyxl. Please install it: pip install openpyxl")
            return redirect('dashboard:patient_prescriptions')
    
    elif format == 'pdf':
        # For PDF, create HTML that can be printed to PDF
        from django.template.loader import render_to_string
        
        context = {
            'prescriptions': prescriptions,
            'patient': request.user,
            'export_date': datetime.now(),
        }
        
        # Reuse the doctor's export template but with patient context
        html_content = render_to_string('dashboard/doctor/prescriptions_export_pdf.html', context)
        
        response = HttpResponse(html_content, content_type='text/html')
        response['Content-Disposition'] = f'attachment; filename="my_prescriptions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"'
        return response
    
    return redirect('dashboard:patient_prescriptions')


@login_required
@patient_required
def patient_profile(request):
    """Handle Patient Profile View & Update"""
    profile, created = PatientProfile.objects.get_or_create(user=request.user)

    if request.method == 'POST':
        
        # ✅ Handle profile picture SEPARATELY (direct model save to S3)
        if 'profile_picture' in request.FILES:
            pic = request.FILES['profile_picture']
            print(f"DEBUG: Uploading file: {pic.name}, Size: {pic.size}")
            
            # Delete old picture from storage (optional but recommended)
            if request.user.profile_picture:
                try:
                    request.user.profile_picture.delete(save=False)
                except Exception as e:
                    print(f"DEBUG: Could not delete old picture: {e}")
            
            # Save new picture directly to user model
            request.user.profile_picture = pic
            request.user.save(update_fields=['profile_picture'])
            print(f"DEBUG: File saved to: {request.user.profile_picture.url}")

        # ✅ Handle User fields directly (simpler than nested serializer for forms)
        user = request.user
        user.first_name = request.POST.get('first_name', user.first_name)
        user.last_name = request.POST.get('last_name', user.last_name)
        user.phone = request.POST.get('phone', user.phone)
        user.gender = request.POST.get('gender', user.gender)
        
        # Handle date_of_birth (can be empty)
        dob = request.POST.get('date_of_birth')
        if dob:
            user.date_of_birth = dob
        
        user.save()

        # ✅ Handle PatientProfile fields
        profile.blood_type = request.POST.get('blood_type', profile.blood_type)
        profile.allergies = request.POST.get('allergies', profile.allergies)
        profile.chronic_conditions = request.POST.get('chronic_conditions', profile.chronic_conditions)
        profile.emergency_contact_name = request.POST.get('emergency_contact_name', profile.emergency_contact_name)
        profile.emergency_contact_phone = request.POST.get('emergency_contact_phone', profile.emergency_contact_phone)
        
        # Handle numeric fields (can be empty)
        height = request.POST.get('height_cm')
        weight = request.POST.get('weight_kg')
        
        if height:
            profile.height_cm = height
        if weight:
            profile.weight_kg = weight
            
        profile.save()

        messages.success(request, 'Profile updated successfully!')
        return redirect('dashboard:patient_profile')

    context = {
        'user': request.user,
        'profile': profile
    }
    return render(request, 'dashboard/patient/profile.html', context)


@login_required
@patient_required
def patient_health_profile(request):
    """View and edit health profile"""
    profile, created = HealthProfile.objects.get_or_create(patient=request.user)
    
    if request.method == 'POST':
        form = HealthProfileForm(request.POST, instance=profile)
        if form.is_valid():
            form.save()
            messages.success(request, 'Health profile updated successfully!')
            return redirect('dashboard:patient_health_profile')
        else:
            for field, errors in form.errors.items():
                messages.error(request, f'{field}: {errors[0]}')
    else:
        form = HealthProfileForm(instance=profile)
    
    context = {
        'form': form,
        'profile': profile,
    }
    return render(request, 'dashboard/patient/health_profile.html', context)


@login_required
@patient_required
def patient_medical_history(request):
    """View and add medical history"""
    history = MedicalHistory.objects.filter(patient=request.user).order_by('-event_date')
    
    if request.method == 'POST':
        form = MedicalHistoryForm(request.POST)
        if form.is_valid():
            entry = form.save(commit=False)
            entry.patient = request.user
            entry.save()
            messages.success(request, 'Medical history entry added!')
            return redirect('dashboard:patient_medical_history')
        else:
            for field, errors in form.errors.items():
                messages.error(request, f'{field}: {errors[0]}')
    else:
        form = MedicalHistoryForm()
    
    context = {
        'history': history,
        'form': form,
    }
    return render(request, 'dashboard/patient/medical_history.html', context)


@login_required
@patient_required
def patient_medical_history_delete(request, pk):
    """Delete a medical history entry"""
    if request.method == 'POST':
        entry = get_object_or_404(MedicalHistory, pk=pk, patient=request.user)
        entry.delete()
        messages.success(request, 'Medical history entry deleted.')
    return redirect('dashboard:patient_medical_history')


@login_required
@patient_required
def patient_medical_documents(request):
    """View and upload medical documents"""
    documents = MedicalDocument.objects.filter(patient=request.user).order_by('-uploaded_at')
    
    if request.method == 'POST':
        form = MedicalDocumentForm(request.POST, request.FILES)
        if form.is_valid():
            doc = form.save(commit=False)
            doc.patient = request.user
            
            # Ensure file_size is set from the uploaded file
            if 'file' in request.FILES:
                uploaded_file = request.FILES['file']
                doc.file_size = uploaded_file.size
                print(f"DEBUG: Setting file_size to {uploaded_file.size} bytes for file: {uploaded_file.name}")
            
            doc.save()
            messages.success(request, 'Document uploaded successfully!')
            return redirect('dashboard:patient_medical_documents')
        else:
            for field, errors in form.errors.items():
                messages.error(request, f'{field}: {errors[0]}')
    else:
        form = MedicalDocumentForm()
    
    context = {
        'documents': documents,
        'form': form,
    }
    return render(request, 'dashboard/patient/medical_documents.html', context)


@login_required
@patient_required
def patient_document_download(request, pk):
    """Download/view a medical document with signed URL"""
    document = get_object_or_404(MedicalDocument, pk=pk, patient=request.user)
    
    # Get the file key (path) from the document
    file_key = document.file.name
    print(f"DEBUG: Generating signed URL for file_key: {file_key}")
    print(f"DEBUG: File object: {document.file}")
    print(f"DEBUG: File storage: {document.file.storage}")
    
    # Generate signed URL manually
    signed_url = generate_signed_url(file_key)
    
    if signed_url:
        print(f"DEBUG: Signed URL generated successfully")
        return HttpResponseRedirect(signed_url)
    else:
        messages.error(request, 'Could not generate download link. Please try again.')
        return redirect('dashboard:patient_medical_documents')


def generate_signed_url(file_key, expiration=3600):
    """Generate a signed URL for Supabase S3 storage"""
    
    try:
        # Get configuration from environment
        endpoint_url = os.getenv('SUPABASE_S3_ENDPOINT_URL')
        access_key = os.getenv('SUPABASE_ACCESS_KEY_ID')
        secret_key = os.getenv('SUPABASE_SECRET_ACCESS_KEY')
        bucket_name = os.getenv('SUPABASE_PRIVATE_BUCKET_NAME', 'medical-records')
        region = os.getenv('SUPABASE_REGION', 'eu-west-2')
        
        if not all([endpoint_url, access_key, secret_key]):
            print(f"Error: Missing Supabase S3 configuration")
            return None
        
        # Create S3 client
        s3_client = boto3.client(
            's3',
            endpoint_url=endpoint_url,
            aws_access_key_id=access_key,
            aws_secret_access_key=secret_key,
            region_name=region,
            config=Config(signature_version='s3v4')
        )
        
        # Ensure file_key doesn't have leading slash (S3 keys shouldn't start with /)
        file_key = file_key.lstrip('/')
        
        # Generate signed URL
        signed_url = s3_client.generate_presigned_url(
            'get_object',
            Params={
                'Bucket': bucket_name,
                'Key': file_key,
            },
            ExpiresIn=expiration
        )
        
        print(f"DEBUG: Generated signed URL for key: {file_key}")
        return signed_url
    
    except Exception as e:
        print(f"Error generating signed URL: {e}")
        import traceback
        traceback.print_exc()
        return None

@login_required
@patient_required
def patient_document_delete(request, pk):
    """Delete a medical document"""
    if request.method == 'POST':
        document = get_object_or_404(MedicalDocument, pk=pk, patient=request.user)
        
        # Delete file from storage
        if document.file:
            document.file.delete(save=False)
        
        document.delete()
        messages.success(request, 'Document deleted successfully.')
    
    return redirect('dashboard:patient_medical_documents')



@patient_required
def patient_notifications(request):
    return render(request, 'dashboard/patient/notifications.html')


# ==========================
# SHARED VIEWS
# ==========================

# def consultations(request):
#     if not request.user.is_authenticated:
#         return redirect('dashboard:login')
    
#     base_template = 'dashboard/base_doctor.html' if request.user.user_type == 'doctor' else 'dashboard/base_patient.html'
#     return render(request, 'dashboard/shared/consultations.html', {'base_template': base_template})

def chat(request):
    if not request.user.is_authenticated:
        return redirect('dashboard:login')
    
    base_template = 'dashboard/base_doctor.html' if request.user.user_type == 'doctor' else 'dashboard/base_patient.html'
    return render(request, 'dashboard/shared/chat.html', {'base_template': base_template})

def video_call(request):
    if not request.user.is_authenticated:
        return redirect('dashboard:login')
    
    base_template = 'dashboard/base_doctor.html' if request.user.user_type == 'doctor' else 'dashboard/base_patient.html'
    return render(request, 'dashboard/shared/video_call.html', {'base_template': base_template})

def voice_call(request):
    if not request.user.is_authenticated:
        return redirect('dashboard:login')
    
    base_template = 'dashboard/base_doctor.html' if request.user.user_type == 'doctor' else 'dashboard/base_patient.html'
    return render(request, 'dashboard/shared/voice_call.html', {'base_template': base_template})

@login_required
def active_encounter(request, appointment_id):
    """
    Active consultation page with video and clinical form.
    - Doctors see full clinical form + video
    - Patients see video + appointment info only
    """
    user = request.user
    is_doctor = getattr(user, 'user_type', None) == 'doctor'
    
    # Determine base template
    base_template = 'dashboard/base_doctor.html' if is_doctor else 'dashboard/base_patient.html'

    # Load appointment with permissions
    if is_doctor:
        appointment = get_object_or_404(
            Appointment.objects.select_related('patient', 'doctor__user', 'doctor__specialization'),
            id=appointment_id,
            doctor=user.doctor_profile,
        )
    else:
        appointment = get_object_or_404(
            Appointment.objects.select_related('patient', 'doctor__user', 'doctor__specialization'),
            id=appointment_id,
            patient=user,
        )

    # Check if appointment can be joined
    if appointment.status not in ['confirmed', 'in_progress', 'pending']:
        messages.error(request, 'This appointment is not available for consultation.')
        if is_doctor:
            return redirect('dashboard:doctor_appointments')
        return redirect('dashboard:patient_appointments')

    # Update status to in_progress if confirmed
    if appointment.status in ['confirmed', 'pending']:
        appointment.status = 'in_progress'
        appointment.save(update_fields=['status'])

    # Ensure video room exists
    if not appointment.video_room_url:
        try:
            appointment.generate_video_room()
            appointment.save(update_fields=['video_room_url', 'video_host_url', 'video_room_id'])
        except Exception as e:
            messages.warning(request, f"Video room unavailable: {e}")

    # Get or create consultation record (for doctors)
    consultation = None
    if is_doctor:
        from consultations.models import Consultation
        consultation, created = Consultation.objects.get_or_create(
            appointment=appointment,
            defaults={
                'chief_complaint': appointment.symptoms or '',
                'symptoms': appointment.symptoms or '',
                'started_at': timezone.now(),
            }
        )
        
        # Update started_at if not set
        if not consultation.started_at:
            consultation.started_at = timezone.now()
            consultation.save(update_fields=['started_at'])

    # Pick host/guest URL based on user type
    raw_room_url = (
        appointment.video_host_url if is_doctor else appointment.video_room_url
    ) or appointment.video_room_url

    # Whereby embed URL with parameters for cleaner interface
    video_embed_url = None
    if raw_room_url:
        display_name = f"Dr.{user.last_name}" if is_doctor else user.first_name
        
        if "?" in raw_room_url:
            video_embed_url = f"{raw_room_url}&embed&displayName={display_name}"
        else:
            video_embed_url = f"{raw_room_url}?embed&displayName={display_name}"

    context = {
        'base_template': base_template,
        'appointment': appointment,
        'patient': appointment.patient,
        'doctor': appointment.doctor,
        'consultation': consultation,
        'video_room_url': raw_room_url,
        'video_embed_url': video_embed_url,
        'is_doctor': is_doctor,
    }

    return render(request, 'dashboard/shared/active_encounter.html', context)


@login_required
def save_encounter_draft(request, appointment_id):
    """
    Save consultation draft without ending the encounter.
    Doctor only.
    """
    user = request.user
    
    if getattr(user, 'user_type', None) != 'doctor':
        messages.error(request, 'Only doctors can save consultation notes.')
        return redirect('dashboard:active_encounter', appointment_id=appointment_id)
    
    appointment = get_object_or_404(
        Appointment,
        id=appointment_id,
        doctor=user.doctor_profile,
    )
    
    if request.method == 'POST':
        from consultations.models import Consultation
        
        consultation, created = Consultation.objects.get_or_create(
            appointment=appointment
        )
        
        # Save consultation data
        consultation.chief_complaint = request.POST.get('chief_complaint', '')
        consultation.symptoms = request.POST.get('symptoms', '')
        consultation.examination_notes = request.POST.get('examination_notes', '')
        consultation.diagnosis = request.POST.get('diagnosis', '')
        consultation.treatment_plan = request.POST.get('treatment_plan', '')
        consultation.notes = request.POST.get('notes', '')
        consultation.private_notes = request.POST.get('private_notes', '')
        
        # Follow-up
        consultation.followup_needed = request.POST.get('followup_needed') == 'on'
        followup_date = request.POST.get('followup_date')
        consultation.followup_date = followup_date if followup_date else None
        consultation.followup_notes = request.POST.get('followup_notes', '')
        
        consultation.save()
        
        messages.success(request, 'Consultation notes saved.')
    
    return redirect('dashboard:active_encounter', appointment_id=appointment_id)


@login_required
def end_encounter(request, appointment_id):
    """
    End the consultation, save all form data, and create prescription.
    Only doctors can end consultations.
    """
    user = request.user
    is_doctor = getattr(user, 'user_type', None) == 'doctor'
    
    if not is_doctor:
        messages.error(request, 'Only doctors can end consultations.')
        return redirect('dashboard:patient_appointments')
    
    appointment = get_object_or_404(
        Appointment,
        id=appointment_id,
        doctor=user.doctor_profile,
    )
    
    if request.method == 'POST':
        from consultations.models import Consultation, Prescription, PrescriptionItem
        
        # Debug: Print all POST data
        print("=" * 50)
        print("POST DATA:")
        for key, value in request.POST.items():
            print(f"  {key}: {value}")
        print("=" * 50)
        print("POST LISTS:")
        print(f"  medicine[]: {request.POST.getlist('medicine[]')}")
        print(f"  dosage[]: {request.POST.getlist('dosage[]')}")
        print(f"  frequency[]: {request.POST.getlist('frequency[]')}")
        print(f"  duration[]: {request.POST.getlist('duration[]')}")
        print(f"  quantity[]: {request.POST.getlist('quantity[]')}")
        print(f"  instruction[]: {request.POST.getlist('instruction[]')}")
        print("=" * 50)
        
        # Get or create consultation
        consultation, created = Consultation.objects.get_or_create(
            appointment=appointment
        )
        
        # Save consultation data
        consultation.chief_complaint = request.POST.get('chief_complaint', '')
        consultation.symptoms = request.POST.get('symptoms', '')
        consultation.examination_notes = request.POST.get('examination_notes', '')
        consultation.diagnosis = request.POST.get('diagnosis', '')
        consultation.treatment_plan = request.POST.get('treatment_plan', '')
        consultation.notes = request.POST.get('notes', '')
        consultation.private_notes = request.POST.get('private_notes', '')
        
        # Follow-up
        consultation.followup_needed = request.POST.get('followup_needed') == 'on'
        followup_date = request.POST.get('followup_date')
        if followup_date:
            consultation.followup_date = followup_date
        else:
            consultation.followup_date = None
        consultation.followup_notes = request.POST.get('followup_notes', '')
        
        # Set end time
        consultation.ended_at = timezone.now()
        
        consultation.save()
        
        # Handle prescription items - try both naming conventions
        medicines = request.POST.getlist('medicine[]') or request.POST.getlist('medicine')
        dosages = request.POST.getlist('dosage[]') or request.POST.getlist('dosage')
        frequencies = request.POST.getlist('frequency[]') or request.POST.getlist('frequency')
        durations = request.POST.getlist('duration[]') or request.POST.getlist('duration')
        quantities = request.POST.getlist('quantity[]') or request.POST.getlist('quantity')
        instructions = request.POST.getlist('instruction[]') or request.POST.getlist('instruction')
        
        print(f"Medicines found: {medicines}")
        
        # Filter out empty medicines
        valid_medicines = [(i, med) for i, med in enumerate(medicines) if med and med.strip()]
        
        print(f"Valid medicines: {valid_medicines}")
        
        # Create prescription if medicines were added
        if valid_medicines:
            prescription = Prescription.objects.create(
                consultation=consultation,
                diagnosis=consultation.diagnosis,
                notes=consultation.notes,
            )
            
            print(f"Created prescription: {prescription.prescription_number}")
            
            for i, medicine in valid_medicines:
                item = PrescriptionItem.objects.create(
                    prescription=prescription,
                    medicine_name=medicine.strip(),
                    dosage=dosages[i].strip() if i < len(dosages) and dosages[i] else '',
                    frequency=frequencies[i] if i < len(frequencies) and frequencies[i] else 'once_daily',
                    duration=durations[i] if i < len(durations) and durations[i] else '7_days',
                    quantity=quantities[i].strip() if i < len(quantities) and quantities[i] else '',
                    instructions=instructions[i].strip() if i < len(instructions) and instructions[i] else '',
                )
                print(f"Created prescription item: {item.medicine_name}")
            
            messages.success(request, f'Prescription {prescription.prescription_number} created.')
        else:
            print("No valid medicines found - no prescription created")
        
        # End the appointment
        appointment.status = 'completed'
        appointment.save(update_fields=['status'])
        
        messages.success(request, f'Consultation {appointment.appointment_number} completed successfully.')
        return redirect('dashboard:doctor_appointments')
    
    return redirect('dashboard:active_encounter', appointment_id=appointment_id)

@login_required
def leave_encounter(request, appointment_id):
    """
    Patient leaves the encounter (doesn't end it, just exits).
    """
    user = request.user
    
    if getattr(user, 'user_type', None) == 'patient':
        messages.info(request, 'You have left the consultation. The doctor may still be adding notes.')
        return redirect('dashboard:patient_appointments')
    
    return redirect('dashboard:active_encounter', appointment_id=appointment_id)

@doctor_required
def consultations_list(request):
    """
    Displays a list of today's video consultations for the doctor.
    Allows them to click 'Join' to enter the Active Encounter.
    """
    return render(request, 'dashboard/doctor/consultations_list.html')
